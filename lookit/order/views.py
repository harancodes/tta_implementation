from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db.models import (
    Value,
    When,
    Case,
    IntegerField,
    Sum,
    Count,
    DecimalField,
)
from django.contrib import messages
from django.db import transaction
from cart.models import Cart
from user.models import Address
from .models import Order, OrderItems, ReturnRequest, Review
from product.models import Variant
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import date, timedelta
from cloudinary.uploader import upload
from core.decorators import admin_required
from .utils import reduce_stock_for_order
from cart.decorators import validate_cart
from wallet.models import Wallet, WalletTransactions
from cart.utils import calculate_cart_summary
from decimal import Decimal
from coupon.utils import (
    reduce_coupon_usage_limit,
    create_coupon_usage_record,
    clear_users_saved_coupon,
    coupon_eligibility_check,
)
from wallet.utils import refund_to_wallet
from django.db.models.functions import Coalesce
from order.utils import render_to_pdf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from coupon.utils import is_coupon_min_purchase_eligible, clear_users_applied_coupon
from urllib.parse import urlencode
from datetime import date, datetime
from .utils import annotate_review_eligibility, restock_inventory
from django.conf import settings


@login_required
@validate_cart
def checkout(request):
    summary = calculate_cart_summary(request.user)
    order_items = summary.get('items')
    price_summary = summary.get('cart_summary')
    applied_coupon = summary.get('applied_coupon')
    
    if applied_coupon:
        purchase_amount = price_summary.get('sub_total') - price_summary.get('offer_discount')
        coupon_eligible = is_coupon_min_purchase_eligible(applied_coupon.code, purchase_amount)
        if not coupon_eligible:
            clear_users_applied_coupon(applied_coupon, request.user)
            messages.error(request,f"This coupon requires a minimum purchase of ₹{applied_coupon.min_purchase_amount}.")
            return redirect('cart')

    # --stock and availability validations--------------------
    for item in order_items:
        error = None
        if not item.product.is_active:
            error = "Product Is Unavailable"
            break
        elif item.quantity < 1 or item.quantity > settings.MAX_CART_QUANTITY:
            # --min and max quantity validation.
            error = f"Cart Quantity Need to be alteast 1 and atmost {settings.MAX_CART_QUANTITY}"
            break
        elif item.available_stock_count == 0:
            error = "Out of stock products in cart."
            break
        elif item.available_stock_count < item.quantity:
            print("stock available: ", item.available_stock_count)
            print("qunatity cart : ", item.quantity)
            error = "Requested quantity not available for some products."
            item.quantity = item.available_stock_count
            item.save()
            break

    if error:
        messages.error(request, error)
        return redirect('cart')

    # ---address list of user, (default address first order)---
    address_list = Address.objects.filter(user=request.user, is_active=True).order_by(
        '-is_default'
    )

    return render(
        request,
        "order/checkout.html",
        {
            "order_items": order_items,
            "price_summary": price_summary,
            "address_list": address_list,
        },
    )


@login_required
@require_POST
@validate_cart
def create_order(request):
    user = request.user
    address_id = request.POST.get('address_id')

    # ---handle error if no address is selected----------------
    if address_id == '0':
        messages.error(request, "PLEASE SELECT AN ADDRESS")
        return redirect('checkout')

    try:
        address = Address.objects.get(id=address_id, user=user)
    except Exception as e:
        messages.error(request, e)
        return redirect('checkout')

    # ---fetch cart summary------------------------------------------------
    summary = calculate_cart_summary(request.user)
    cart_items = summary.get('items')
    order_summary = summary.get('cart_summary')
    total_items = cart_items.count()
    applied_coupon = (
        summary.get('applied_coupon') if summary.get('applied_coupon') else None
    )
    order = None
    try:
        with transaction.atomic():

            order = Order.objects.create(
                user=user,
                address=address,
                total_items=total_items,  # number of products
                sub_total=order_summary.get('sub_total'),
                delivery_total=order_summary.get('delivery_fee'),
                discount_total=order_summary.get('offer_discount'),
                coupon_applied=applied_coupon,
                coupon_discount_total=order_summary.get('coupon_discount'),
                grand_total=order_summary.get('grand_total'),
            )
            for item in cart_items:
                coupon_discount = None
                final_amount = item.sub_total
                delivery_fee = Decimal(50.00)
                
                if applied_coupon:
                    # Calculate proportional coupon discount for each product
                    price_after_discount = (
                        item.sub_total_per_product - item.discount_subtotal
                    )
                    total_order_value = order_summary.get(
                        'sub_total'
                    ) - order_summary.get('offer_discount')
                    coupon_discount = Decimal(
                        (price_after_discount / total_order_value)
                        * order_summary.get('coupon_discount')
                    )
                    final_amount -= coupon_discount
                    
                #add delivery charge
                final_amount += delivery_fee

                OrderItems.objects.create(
                    order=order,
                    variant=item.variant,
                    quantity=item.quantity,
                    unit_price=item.variant.product.price,
                    sub_total=item.sub_total_per_product,
                    delivery_fee=delivery_fee,
                    discount_amount=item.discount_subtotal,
                    coupon_discount_amount=coupon_discount,
                    total=final_amount,  # final line total (after deducted discount amount and coupon amount)
                )

    except Exception as e:
        messages.error(request, e)
        print("Error: ", e)
        return redirect('checkout')

    return redirect('payment-page', order_uuid=order.uuid)


@login_required
@validate_cart
def payment_page(request, order_uuid):
    order = Order.objects.get(uuid=order_uuid)
    wallet, created = Wallet.objects.get_or_create(user=request.user)
    wallet_balance_after_payment = wallet.balance - order.grand_total
    
    address = order.address
    # estimated delivery date - after 7 days from today
    today = timezone.now().date()
    estimated_delivery = today + timedelta(days=7)
    return render(
        request,
        "order/payment.html/",
        {
            "order": order,
            "address": address,
            "wallet": wallet,
            "wallet_balance_after_payment": wallet_balance_after_payment,
            "estimated_delivery": estimated_delivery,
            "cod_min_order_amount": settings.COD_MIN_ORDER_AMOUNT
        },
    )


@login_required
@validate_cart
def place_order(request, order_uuid):
    order = Order.objects.get(uuid=order_uuid)
    order_id = order.id
    if request.method == "POST":
        payment_method = request.POST.get('payment_method')

        # ---check if payment method is valid
        if payment_method not in Order.PaymentMethod.values:
            messages.error(request, "Invalid Payment Method")
            return redirect('payment-page', order_id=order_id)

        try:
            with transaction.atomic():
                Order.objects.filter(id=order_id).update(payment_method=payment_method)
                if payment_method == 'WALLET':
                    # --deduct amount from wallet--
                    amount = order.grand_total
                    wallet = Wallet.objects.get(user=request.user)
                    wallet.balance -= amount
                    wallet.save()
                    # ---create wallet transaction--------------------
                    WalletTransactions.objects.create(
                        wallet=wallet,
                        order=order,
                        amount=amount,
                        transaction_type=WalletTransactions.TransactionType.DEBIT,
                        label="Paid on Shopping",
                        txn_source=WalletTransactions.TransactionSource.SHOPPING,
                    )
                    # ---place order------------------------------------
                    OrderItems.objects.filter(order_id=order_id).update(
                        order_status=OrderItems.OrderStatus.PLACED,
                        payment_status=OrderItems.PaymentStatus.PAID,
                        placed_at=timezone.now(),
                    )
                elif payment_method == 'COD':
                    
                    #only orders less than or equal to cod_max_amount_limit can be placed with cod
                    amount = order.grand_total
                    if amount > settings.COD_MIN_ORDER_AMOUNT:
                        messages.error(request, f"COD is available only for orders up to ₹{settings.COD_MIN_ORDER_AMOUNT}.")
                        return redirect('payment-page', order_uuid=order_uuid)
                    
                    OrderItems.objects.filter(order_id=order_id).update(
                        order_status=OrderItems.OrderStatus.PLACED,
                        payment_status=OrderItems.PaymentStatus.COD,
                        placed_at=timezone.now(),
                    )

                # --handle stock count of the product---
                reduce_stock_for_order(order_id)

                # --reduce coupon usage limit if applied any---
                order = Order.objects.get(id=order_id)
                if order.coupon_applied:
                    # check if user already used the coupon
                    eligible = coupon_eligibility_check(
                        order.coupon_applied.code, request.user
                    )
                    if not eligible:
                        messages.error(request, "You already used coupon once")
                        return redirect('payment-page', order_uuid=order_uuid)

                    reduce_coupon_usage_limit(order.coupon_applied.code)
                    create_coupon_usage_record(order.coupon_applied, request.user)
                    clear_users_saved_coupon(order.coupon_applied, request.user)

                messages.success(request, "ORDER PLACED SUCCESSFULLY")
                if payment_method == 'WALLET':
                    messages.success(
                        request, f"{order.grand_total} deducted from wallet"
                    )

        except Exception as e:
            messages.error(request, e)
            return redirect('payment-page', order_uuid=order_uuid)

        # ---redirect to order success page---------------
        return redirect('order-success', order_uuid=order_uuid)


@login_required
def order_success_page(request, order_uuid):

    # ---empty the cart of user--------------------
    Cart.objects.filter(user=request.user).delete()

    order = Order.objects.get(uuid=order_uuid)
    order_items = OrderItems.objects.filter(order__uuid=order_uuid)
    address = order.address
    print(order.payment_method)
    return render(
        request,
        "order/order_success.html",
        {'order': order, "address": address, "order_items": order_items},
    )


@login_required
def payment_failure_page(request, order_uuid):
    order = Order.objects.get(uuid=order_uuid)
    return render(request, 'order/payment_failed.html', {'order': order})


@login_required
def my_orders(request):
    orders = (
        OrderItems.objects.filter(order__user=request.user)
        .exclude(order_status='INITIATED')
        .annotate(
            status_priority=Case(
                    When(order_status='OUT_FOR_DELIVERY', then=Value(1)),
                    When(order_status='SHIPPED', then=Value(2)),
                    When(order_status='PLACED', then=Value(3)),
                    When(order_status='DELIVERED', then=Value(4)),
                    When(order_status='RETURNED', then=Value(5)),
                    When(order_status='CANCELLED', then=Value(6)),
                    default=Value(7),
                    output_field=IntegerField(),
                )
        )
        .order_by('status_priority', '-created_at')
    )
    # --search functionality----------
    search = request.GET.get('search')
    if search:
        orders = orders.filter(variant__product__name__icontains=search)
        
    #annotate a field for determining whether to show write a review btn
    orders = annotate_review_eligibility(request.user, orders)
    
    return render(request, "order/my_orders.html", {"orders": orders})


@login_required
def track_order(request, order_uuid):
    order_item = OrderItems.objects.get(uuid=order_uuid)
    delivery_address = order_item.order.address

    # --for checking return eligibility if delivered
    seven_days_ago = timezone.now() - timedelta(days=7)
    delivered_date = order_item.delivered_at
    eligible_for_return = False
    if delivered_date:
        if delivered_date > seven_days_ago:
            eligible_for_return = True  # 7 days not passed

    return render(
        request,
        "order/track_order.html",
        {
            "order": order_item,
            "address": delivery_address,
            "eligible_for_return": eligible_for_return,
        },
    )


from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from io import BytesIO


def download_invoice_pdf(request, order_uuid):
    order_item = OrderItems.objects.get(uuid=order_uuid)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()

    # Custom Title Style
    header_style = ParagraphStyle(
        "HeaderStyle",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#9a4d72"),
        alignment=1,  # center
        spaceAfter=10,
        fontName="Helvetica-Bold",
    )

    section_title = ParagraphStyle(
        "SectionTitle",
        fontSize=13,
        textColor=colors.HexColor("#444"),
        spaceBefore=12,
        spaceAfter=6,
        fontName="Helvetica-Bold",
    )

    normal_bold = ParagraphStyle(
        "NormalBold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
    )

    elements = []

    # -------------------------------------------
    # 🔶 BRAND HEADER
    # -------------------------------------------
    elements.append(Paragraph("LookIt Private Limited", header_style))
    elements.append(Paragraph("GSTIN: 22AAAAA0000A1Z5", styles["Normal"]))  # Optional
    elements.append(Paragraph("support@lookit.com | www.lookit.com", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>Invoice</b>", section_title))
    elements.append(Spacer(1, 5))

    # -------------------------------------------
    # 🔶 ORDER INFORMATION BLOCK
    # -------------------------------------------
    order_info = [
        ["Invoice No:", order_item.uuid],
        ["Order Date:", order_item.placed_at.strftime("%d %b %Y")],
        [
            "Delivered Date:",
            (
                order_item.delivered_at.strftime("%d %b %Y")
                if order_item.delivered_at
                else "-"
            ),
        ],
        ["Customer Name:", order_item.order.user.full_name],
        ["Payment Method:", order_item.order.payment_method],
    ]

    info_table = Table(order_info, colWidths=[110, 360])
    info_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elements.append(info_table)
    elements.append(Spacer(1, 18))

    # -------------------------------------------
    # 🔶 PRODUCT ITEM TABLE
    # -------------------------------------------
    elements.append(Paragraph("<b>Order Item Details</b>", section_title))

    items_data = [["Product Name", "Qty", "Price", "Subtotal"]]

    items_data.append(
        [
            order_item.product.name,
            str(order_item.quantity),
            f"Rs {order_item.unit_price}",
            f"Rs {order_item.sub_total}",
        ]
    )

    items_table = Table(items_data, colWidths=[260, 60, 80, 90])
    items_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f2f2")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 1, colors.black),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )

    elements.append(items_table)
    elements.append(Spacer(1, 20))

    # -------------------------------------------
    # 🔶 TOTAL SUMMARY
    # -------------------------------------------
    elements.append(Paragraph("<b>Payment Summary</b>", section_title))

    totals_data = []

    # Subtotal (always)
    totals_data.append(
        ["Subtotal:", f"Rs {order_item.sub_total}"]
    )

    # Product discount (only if applied)
    if order_item.discount_amount > 0:
        totals_data.append(
            ["Product Discount:", f"-Rs {order_item.discount_amount}"]
        )

    # Coupon discount (only if applied)
    coupon = order_item.order.coupon_applied
    if coupon and order_item.coupon_discount_amount > 0:
        totals_data.append(
            [f"Coupon Discount ({coupon.code}):", f"-Rs {order_item.coupon_discount_amount}"]
        )

    # Delivery fee (always)
    totals_data.append(
        ["Delivery Fee:", f"Rs {order_item.delivery_fee}"]
    )

    # Grand total (always last)
    totals_data.append(
        ["Grand Total:", f"Rs {order_item.total}"]
    )


    totals_table = Table(totals_data, colWidths=[150, 120])
    totals_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 2), "Helvetica"),
                ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    elements.append(totals_table)
    elements.append(Spacer(1, 20))

    # -------------------------------------------
    # 🔶 FOOTER
    # -------------------------------------------
    elements.append(Spacer(1, 30))
    elements.append(
        Paragraph("<b>Thank you for shopping with LookIt!</b>", styles["Italic"])
    )
    elements.append(
        Paragraph(
            "This is a system-generated invoice and does not require a signature.",
            styles["Normal"],
        )
    )

    # Build PDF
    doc.build(elements)

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response['Content-Disposition'] = (
        f'inline; filename="Invoice-{order_item.uuid}.pdf"'
    )
    return response


@login_required
def cancel_order(request, order_item_uuid):
    try:
        order_item = OrderItems.objects.get(uuid=order_item_uuid)

        if order_item.order.user == request.user:
            with transaction.atomic():
                order_item.order_status = 'CANCELLED'
                order_item.cancelled_at = timezone.now()
                order_item.save()
                restock_inventory(order_item_uuid)

                # refund money back to wallet instantly (except cod)
                if order_item.order.payment_method != Order.PaymentMethod.COD:
                    refund_amount = order_item.total
                    refund_to_wallet(request.user, refund_amount, order_item)
                    messages.success(request, "Order Cancelled.")
                    messages.success(
                        request, f"₹{refund_amount} has been credited to your wallet."
                    )
                else:
                    messages.success(request, "Order Cancelled.")

        else:
            messages.error(request, "Unauthorized Access")
    except Exception as e:
        print("Error on cacellation: ",e)
        messages.error(request, "Something went wrong!")
    return redirect('track-order', order_uuid=order_item_uuid)


@login_required
def return_request_form(request, order_uuid):
    if request.method == "POST":

        order_item = OrderItems.objects.get(uuid=order_uuid)

        # --for checking return eligibility if delivered
        seven_days_ago = timezone.now() - timedelta(days=7)
        delivered_date = order_item.delivered_at
        eligible_for_return = False
        if delivered_date:
            if delivered_date > seven_days_ago:
                eligible_for_return = True  # 7 days not passed

        if not eligible_for_return:
            messages.error(request, "Not Eligible For Return")
            return redirect('return-request-form', order_uuid=order_uuid)

        reason = request.POST.get('reason')
        comments = request.POST.get('comments')

        images = request.FILES.getlist('product_images')
        print(request.FILES)
        print("images", images)

        # ---fetch pick up address--------------------------------
        pickup_address_id = request.POST.get('pickup_address_id')
        pickup_address = None
        if pickup_address_id:
            pickup_address = Address.objects.get(id=pickup_address_id)
        else:
            messages.error(request, "Please Select A Pick Up Address")
            return redirect('return-request-form', order_uuid=order_uuid)

        try:
            with transaction.atomic():
                return_request = ReturnRequest.objects.create(
                    order_item=order_item,
                    reason=reason,
                    comments=comments,
                    pickup_address=pickup_address,
                    amount_paid=order_item.total,
                )
                # Assign images to available fields (up to 3)
                if len(images) > 0:
                    result = upload(
                        images[0],
                        folder=f"return_product_images/{request.user.id}",
                        transformation=[
                            {
                                'width': 500,
                                'height': 500,
                                'crop': 'fill',
                                'gravity': 'face',
                            },
                            {
                                'quality': 'auto',
                                'fetch_format': 'auto',
                            },
                        ],
                    )
                    return_request.product_image1 = result['secure_url']
                if len(images) > 1:
                    result = upload(
                        images[1],
                        folder=f"return_product_images/{request.user.id}",
                        transformation=[
                            {
                                'width': 500,
                                'height': 500,
                                'crop': 'fill',
                                'gravity': 'face',
                            },
                            {
                                'quality': 'auto',
                                'fetch_format': 'auto',
                            },
                        ],
                    )
                    return_request.product_image2 = result['secure_url']
                if len(images) > 2:
                    result = upload(
                        images[2],
                        folder=f"return_product_images/{request.user.id}",
                        transformation=[
                            {
                                'width': 500,
                                'height': 500,
                                'crop': 'fill',
                                'gravity': 'face',
                            },
                            {
                                'quality': 'auto',
                                'fetch_format': 'auto',
                            },
                        ],
                    )
                    return_request.product_image3 = result['secure_url']
                return_request.save()

                messages.success(request, "Return Request Submitted Successfully")

        except Exception as e:
            print(e)
            messages.error(request, e)
            return redirect('return-request-form', order_uuid=order_uuid)

        return redirect('track_return_request', order_uuid=order_uuid)

    order = OrderItems.objects.get(uuid=order_uuid)
    address_list = Address.objects.filter(user=request.user, is_active=True).order_by(
        '-is_default', '-created_at'
    )
    return render(
        request,
        "order/return_request_form.html",
        {"order": order, "address_list": address_list},
    )


@login_required
def track_return_request(request, order_uuid):
    order_item = OrderItems.objects.get(uuid=order_uuid)
    return_request = order_item.return_request

    delivery_address = order_item.order.address
    return render(
        request,
        "order/track_return.html",
        {
            "order": order_item,
            "address": delivery_address,
            "return_request": return_request,
        },
    )
     
@login_required
def write_review(request, order_uuid):
    order = OrderItems.objects.get(uuid=order_uuid)

    if request.method == "POST":
        rating = request.POST.get('rating')
        review = request.POST.get('review')
        print(rating, review)
        try:
            Review.objects.create(product=order.product, user=request.user, rating=rating, review=review)
            messages.success(request, "Review submitted")
            return redirect('my-orders')
        except Exception as e:
            print("Error on creating review: ",e)
            messages.error(request, "Something went wrong")
            return redirect('write-review', order_uuid=order_uuid)
        
    return render(request, "order/write_review.html",{"order":order})


"""
---------ADMIN SIDE------------------------------------------------------------------------- 
"""


@admin_required
def admin_list_orders(request):
    order_items = OrderItems.objects.exclude(order_status='INITIATED').order_by(
        '-placed_at'
    )

    search = request.GET.get('search')
    payment_method = request.GET.get('payment_method')
    payment_status = request.GET.get('payment_status')
    order_status = request.GET.get('order_status')
    date_range = request.GET.get('date_range')

    if search:
        order_items = order_items.filter(
            Q(uuid=search)
            | Q(order__user__full_name__icontains=search)
            | Q(variant__product__name__icontains=search)
        )

    if payment_method:
        order_items = order_items.filter(order__payment_method=payment_method.upper())

    if payment_status:
        order_items = order_items.filter(payment_status=payment_status.upper())

    if order_status:
        order_items = order_items.filter(order_status=order_status.upper())

        # ---- DATE filter ----
    if date_range:
        today = date.today()

        if date_range == 'today':
            order_items = order_items.filter(placed_at__date=today)

        elif date_range == 'week':
            start_of_week = today - timedelta(days=today.weekday())
            order_items = order_items.filter(placed_at__date__gte=start_of_week)

        elif date_range == 'month':
            order_items = order_items.filter(
                placed_at__year=today.year, placed_at__month=today.month
            )
        elif date_range == 'year':
            order_items = order_items.filter(placed_at__year=today.year)

    # pagination
    paginator = Paginator(order_items, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "order/admin/list.html", {"page_obj": page_obj})


@admin_required
def admin_order_details(request, order_item_uuid):
    order_item = OrderItems.objects.get(uuid=order_item_uuid)
    customer = order_item.order.user
    address = order_item.order.address
    return render(
        request,
        "order/admin/order_details.html",
        {"order": order_item, "customer": customer, "address": address},
    )


@admin_required
def admin_update_delivery_status(request, order_item_uuid):
    if request.method == "POST":
        order_status = request.POST.get("order_status")
        try:
            order_item = OrderItems.objects.get(uuid=order_item_uuid)
            if order_status.upper() in OrderItems.OrderStatus.values:
                order_item.order_status = order_status.upper()

                if order_status == 'placed':
                    order_item.placed_at = timezone.now()
                    order_item.shipped_at = None
                    order_item.out_for_delivery_at = None
                    order_item.delivered_at = None
                elif order_status == 'shipped':
                    order_item.shipped_at = timezone.now()
                    order_item.out_for_delivery_at = None
                    order_item.delivered_at = None
                elif order_status == 'out_for_delivery':
                    order_item.out_for_delivery_at = timezone.now()
                    order_item.delivered_at = None
                elif order_status == 'delivered':
                    order_item.delivered_at = timezone.now()
                elif order_status == 'cancelled':
                    order_item.cancelled_at = timezone.now()

                order_item.save()
                messages.success(request, "ORDER STATUS UPDATED SUCCESSFULLY")
            else:
                messages.error(request, "INVALID ORDER STATUS")
        except Exception as e:
            messages.error(request, e)

    return redirect('admin-order-details', order_item_uuid=order_item_uuid)


@admin_required
def admin_list_return_requests(request):
    return_request_list = (
        ReturnRequest.objects.all()
        .select_related('order_item')
        .order_by("-request_date")
    )

    # ---search filter-----------------
    search = request.GET.get('search')
    print(search, "hello")
    if search:
        return_request_list = return_request_list.filter(
            Q(uuid=search)
            | Q(order_item__variant__product__name__icontains=search)
            | Q(order_item__order__user__full_name__icontains=search)
        )

    return_status = request.GET.get("return_status")
    if return_status:
        return_request_list = return_request_list.filter(status=return_status.upper())

    # ---- DATE filter ----
    date_range = request.GET.get('date_range')
    if date_range:
        today = date.today()

        if date_range == 'today':
            return_request_list = return_request_list.filter(request_date__date=today)

        elif date_range == 'week':
            start_of_week = today - timedelta(days=today.weekday())
            return_request_list = return_request_list.filter(
                request_date__date__gte=start_of_week
            )

        elif date_range == 'month':
            return_request_list = return_request_list.filter(
                request_date__year=today.year, request_date__month=today.month
            )
        elif date_range == 'year':
            return_request_list = return_request_list.filter(
                request_date__year=today.year
            )

    paginator = Paginator(return_request_list, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(
        request, 'order/admin/return_request_list.html', {"page_obj": page_obj}
    )


@admin_required
def admin_return_details(request, return_request_uuid):
    return_request = ReturnRequest.objects.get(uuid=return_request_uuid)
    order_item = return_request.order_item
    customer = order_item.order.user
    address = order_item.order.address

    return render(
        request,
        "order/admin/return_details.html",
        {
            "return_request": return_request,
            "order": order_item,
            "customer": customer,
            "address": address,
        },
    )


@admin_required
def admin_update_return_status(request, return_request_uuid):
    if request.method == "POST":
        return_status = request.POST.get("return_status")
        pickup_date = request.POST.get("pickup_date")

        if return_status in ReturnRequest.ReturnStatus.values:
            try:
                return_request = ReturnRequest.objects.get(uuid=return_request_uuid)
                return_request.status = return_status

                # Set relevant timestamp based on new status
                if return_status == ReturnRequest.ReturnStatus.APPROVED:
                    return_request.approved_at = timezone.now()
                    return_request.rejected_at = None
                    return_request.pickup_scheduled_on = None
                    return_request.pickup_scheduled_for = None
                    return_request.pickedup_at = None
                    return_request.refunded_at = None
                elif return_status == ReturnRequest.ReturnStatus.REJECTED:
                    return_request.rejected_at = timezone.now()
                    return_request.approved_at = None
                    return_request.pickup_scheduled_on = None
                    return_request.pickup_scheduled_for = None
                    return_request.pickedup_at = None
                    return_request.refunded_at = None
                elif return_status == ReturnRequest.ReturnStatus.PICKUP_SCHEDULED:
                    if pickup_date:
                        return_request.pickup_scheduled_on = timezone.now()
                        return_request.pickup_scheduled_for = pickup_date
                        return_request.pickedup_at = None
                        return_request.refunded_at = None
                    else:
                        messages.error(request, "Please Provide Pickup Date.")
                elif return_status == ReturnRequest.ReturnStatus.PICKED_UP:
                    return_request.pickedup_at = timezone.now()
                    return_request.refunded_at = None
                    # --update order status ----------
                    order = return_request.order_item
                    order.order_status = order.OrderStatus.RETURNED
                    order.save()
                    # --handle stock------------------
                    restock_inventory(return_request.order_item.uuid)

                elif return_status == ReturnRequest.ReturnStatus.REFUNDED:
                    order_item = return_request.order_item

                    # refund money back to wallet instantly
                    refund_amount = order_item.total
                    refund_to_wallet(order_item.order.user, refund_amount, order_item)
                    messages.success(
                        request, f"₹{refund_amount} has been credited to users wallet."
                    )

                    return_request.refunded_at = timezone.now()
                    # --update order status ----------
                    order_item.order_status = order_item.OrderStatus.REFUNDED
                    order_item.save()

                return_request.save()
                messages.success(request, "Return status updated successfully.")
            except Exception as e:
                print("Error : ", e)
                messages.error(request, "Something went wrong")

        else:
            messages.error(request, "Invalid return status.")

    return redirect("admin-return-details", return_request_uuid=return_request_uuid)


@admin_required
def sales_report(request):
    orders = OrderItems.objects.filter(
        order_status=OrderItems.OrderStatus.DELIVERED
    ).order_by('-placed_at')

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    

    if from_date and to_date:
        from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

        #from date need to be before start date
        if from_date > to_date:
            messages.error(
                request,
                "From date must be earlier than or equal to To date."
            )
            return redirect(request.path)
        
        orders = orders.filter(placed_at__date__range=[from_date, to_date])
        
    else:
        #default filter today
        params = {
            'from_date': date.today().isoformat(),
            'to_date': date.today().isoformat(),
        }
        return redirect(f"{request.path}?{urlencode(params)}")
        

    report_summary = orders.aggregate(
        total_order_count=Count('id'),
        total_order_value=Coalesce(Sum('sub_total'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
        total_offer_discounts=Coalesce(Sum('discount_amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
        total_coupon_discounts=Coalesce(Sum('coupon_discount_amount'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
        realized_revenue=Coalesce(Sum('total'), Value(Decimal('0.00')), output_field=DecimalField(max_digits=10, decimal_places=2)),
    )
    report_summary['total_discounts'] = (
        report_summary['total_offer_discounts']
        + report_summary['total_coupon_discounts']
    )

    # pagination
    paginator = Paginator(orders, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Preserve query params for pagination
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    preserved_query = query_params.urlencode()

    return render(
        request,
        "order/admin/sales_report.html",
        {
            "page_obj": page_obj,
            "preserved_query": preserved_query,
            "report_summary": report_summary,
        },
    )



def download_sales_report_pdf(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get("end_date")
    print(start_date, end_date)
    
    sales = OrderItems.objects.filter(
        order_status=OrderItems.OrderStatus.DELIVERED
    ).order_by('-created_at')
    today = timezone.now().date()
    
    if start_date and end_date:
        sales = sales.filter(created_at__date__range=[start_date, end_date])
    
    report_summary = sales.aggregate(
        total_order_count=Count('id'),
        total_order_value=Sum('sub_total'),
        total_offer_discounts=Sum('discount_amount'),
        total_coupon_discounts=Coalesce(Sum('coupon_discount_amount'), Decimal('0.00'), output_field=DecimalField(max_digits=10, decimal_places=2)),
        realized_revenue=Sum('total'),
    )
    report_summary['total_discounts'] = (
        report_summary['total_offer_discounts']
        + report_summary['total_coupon_discounts']
    )

    context = {
        "sales": sales,
        "report_summary": report_summary,
        "generated_on": today,
        "start_date":start_date,
        "end_date":end_date,
    }

    pdf = render_to_pdf("order/admin/sales_report_exports/sales_report_pdf.html", context)

    if not pdf:
        return HttpResponse("Error generating PDF", status=500)

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="sales_report.pdf"'
    return response

def export_sales_report_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get("end_date")
    
    
    sales = OrderItems.objects.filter(
        order_status=OrderItems.OrderStatus.DELIVERED
    ).order_by('-created_at')
    
    if start_date and end_date:
        sales = sales.filter(created_at__date__range=[start_date, end_date])
    
    report_summary = sales.aggregate(
        total_order_count=Count('id'),
        total_order_value=Sum('sub_total'),
        total_offer_discounts=Sum('discount_amount'),
        total_coupon_discounts=Coalesce(Sum('coupon_discount_amount'), Decimal('0.00'), output_field=DecimalField(max_digits=10, decimal_places=2)),
        realized_revenue=Sum('total'),
    )
    report_summary['total_discounts'] = (
        report_summary['total_offer_discounts']
        + report_summary['total_coupon_discounts']
    )
    
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # ---------- Styles ----------
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")

    header_fill = PatternFill(
        start_color="E2E8F0",
        end_color="E2E8F0",
        fill_type="solid"
    )

    # ---------- Title ----------
    ws.merge_cells("A1:H1")
    ws["A1"] = "LookIt - Sales Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Period: {start_date} - {end_date}"
    ws["A2"].alignment = center_align

    ws.append([])

    # ---------- Summary ----------
    summary_headers = [
        "Total Orders",
        "Total Value",
        "Realized Revenue",
        "Discounts",
    ]

    summary_values = [
        report_summary.get('total_order_count'),
        report_summary.get('total_order_value'),
        report_summary.get('realized_revenue'),
        report_summary.get('total_discounts'),
    ]

    ws.append(summary_headers)
    ws.append(summary_values)

    for col in range(1, 5):
        ws.cell(row=4, column=col).font = header_font
        ws.cell(row=4, column=col).fill = header_fill
        ws.cell(row=4, column=col).alignment = center_align
        ws.cell(row=5, column=col).alignment = center_align

    ws.append([])

    # ---------- Table Header ----------
    headers = [
        "Order ID",
        "Date",
        "Customer",
        "Product",
        "Payment",
        "Coupon",
        "Discount",
        "Total",
    ]

    ws.append(headers)
    header_row = ws.max_row

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # ---------- Data Rows ----------
    for item in sales:
        ws.append([
            item.uuid,
            item.created_at.strftime("%d %b %Y"),
            item.order.user.full_name,
            item.product.name,
            "Razorpay" if item.order.payment_method == "ONLINE_PAYMENT" else item.order.payment_method,
            item.order.coupon_applied.code if item.order.coupon_applied else "-",
            float(item.discount_amount or 0),
            float(item.total),
        ])

    # ---------- Align Amount Columns ----------
    for row in ws.iter_rows(min_row=header_row + 1):
        row[6].alignment = right_align  # Discount
        row[7].alignment = right_align  # Total

    # ---------- Auto column width ----------
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

    # ---------- Response ----------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)
    return response
